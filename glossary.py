from utils import *

class glossary:
    data = {}
    def __init__(self, filename):
        from openpyxl import load_workbook
        sheet = load_workbook(filename)
        sheet_name = sheet.sheetnames[0]
        sheet = sheet[sheet_name]

        for i,row in enumerate(sheet.iter_rows()):
            if i == 0:
                continue
            term, definition = row[0].value, row[1].value
            if term and definition:
                self.data[term] = definition

        print(self.data)

    #传入名词解释表，传出对应的assistant
    def get_definition(self, terms):
        answer = []
        for term in terms:
            definition = self.data[term]
            answer.append(interact_prompt(definition))
        return answer